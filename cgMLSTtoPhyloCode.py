#!/usr/bin/python3
# -*- coding:utf-8 -*-

import argparse
import os
import sys
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import subprocess
import tempfile
import shutil
from pathlib import Path
import openpyxl

#########################################################################################
################### Read cgMLST profiles and convert them to PhyloCode###################
#########################################################################################

def read_cgMLST(cgMLST_locus_file, pubMLST_files, Enterobase_cgMLST):
	# ---------- Load correspondence table ----------
	print('cgMLST_locus_file')
	corr_path = cgMLST_locus_file
	corr_df = pd.read_csv(corr_path, sep='\t', header=None, names=["valid_name", "alt_name"])
	corr_mapping = corr_df.dropna().set_index("alt_name")["valid_name"].to_dict()
	valid_names = corr_df["valid_name"].dropna().unique().tolist()

	# ---------- Initialize pubMLST DataFrame with valid loci as index ----------
	print('Initialize pubMLST DataFrame with valid loci as index')
	df_pubMLST = pd.DataFrame(index=pd.Index(valid_names, name='Label'))

	# ---------- Load and align pubMLST files ----------
	print('Load and align pubMLST files')

	if pubMLST_files != 'None':
		for file_path in pubMLST_files:
			sample_name = os.path.splitext(os.path.basename(file_path))[0]
			print(sample_name)
			
			# Read from Excel instead of CSV/TSV
			df = pd.read_excel(file_path, usecols=[0, 1], header=0)
			df.columns = ['Label', sample_name]

			# Replace "-" with -1 and convert to integers
			df[sample_name] = df[sample_name].replace("-", -1).astype(int)

			# Map alternative names to valid names and drop unmapped ones
			df['Label'] = df['Label'].replace(corr_mapping)
			df = df[df['Label'].isin(valid_names)]
			df = df.groupby("Label")[sample_name].min()  # ensure unique index

			# Add aligned column to the master DataFrame
			df_pubMLST[sample_name] = df_pubMLST.index.to_series().map(df)

	# ---------- Load Enterobase_cgMLST ----------
	print(Enterobase_cgMLST)
	if Enterobase_cgMLST and Enterobase_cgMLST != 'None':
		print(f"Loading Enterobase_cgMLST from: {Enterobase_cgMLST}")
		
		# Load the file
		raw_df = pd.read_csv(Enterobase_cgMLST, sep='\t', header=None)

		# Extract locus labels and sample data
		loci = raw_df.iloc[0, 2:].tolist()
		data = raw_df.iloc[1:].copy()
		sample_names = data.iloc[:, 0].tolist()
		data_values = data.iloc[:, 2:].reset_index(drop=True)

		# Replace "-" with -1
		data_values = data_values.replace("-", -1)

		# Construct DataFrame
		df_enterobase = pd.DataFrame(data_values.transpose().values, index=loci, columns=sample_names)
		df_enterobase = df_enterobase.astype(int)

		# Map alternative locus names to valid ones
		df_enterobase.index = df_enterobase.index.to_series().replace(corr_mapping)

		# Filter only valid locus names
		df_enterobase = df_enterobase[df_enterobase.index.isin(valid_names)]

		# Group in case of duplicates
		df_enterobase = df_enterobase.groupby(df_enterobase.index).min(numeric_only=True)

	# ---------- Combine and sort ----------
	print('Combine and sort')

	# Initialize empty DataFrame with valid index
	combined_df = pd.DataFrame(index=valid_names)

	# Add pubMLST data if available
	if pubMLST_files != 'None':
		combined_df = combined_df.join(df_pubMLST, how='left')

	# Add Enterobase data if available
	if Enterobase_cgMLST != 'None':
		combined_df = combined_df.join(df_enterobase, how='left')

	# Replace any remaining NaNs with -1
	combined_df = combined_df.fillna(-1)

	# Return result
	return(combined_df)

############################ Nearest allele search begin ############################
def check_command_exists(command_name):
	"""Check whether an external command exists in PATH."""
	if shutil.which(command_name) is None:
		print(f"Error: '{command_name}' was not found in PATH.", file=sys.stderr)
		sys.exit(1)

def make_blast_database(db_fasta, db_prefix):
	"""Create a BLAST nucleotide database from a FASTA file."""
	cmd = [
		"makeblastdb",
		"-in", str(db_fasta),
		"-dbtype", "nucl",
		"-out", str(db_prefix),
	]

	try:
		subprocess.run(cmd, check=True, capture_output=True, text=True)
	except subprocess.CalledProcessError as e:
		print("Error: makeblastdb failed.", file=sys.stderr)
		if e.stderr:
			print(e.stderr, file=sys.stderr)
		sys.exit(e.returncode)

def run_blastn(query_fasta, db_prefix):
	"""Run blastn with outfmt 6 against the specified BLAST database."""
	cmd = [
		"blastn",
		"-query", str(query_fasta),
		"-db", str(db_prefix),
		"-outfmt", "6",
		"-qcov_hsp_perc", "99",
		"-perc_identity", "99"
	]
	try:
		result = subprocess.run(cmd, check=True, capture_output=True, text=True)
	except subprocess.CalledProcessError as e:
		print(f"Error: blastn failed for query: {query_fasta}", file=sys.stderr)
		if e.stderr:
			print(e.stderr, file=sys.stderr)
		return None

	return result.stdout


def get_top_hit(blast_output):
	"""
	Return the top hit line based on the last column (bit score).
	Returns None if there are no valid hits.
	"""
	top_hit = None
	top_bitscore = None

	for line in blast_output.splitlines():
		line = line.strip()
		if not line or line.startswith("#"):
			continue

		cols = line.split("\t")
		cols = [""] + cols

		try:
			bitscore = float(cols[-1])
		except (ValueError, IndexError):
			print(f"Warning: skipped invalid BLAST line: {line}", file=sys.stderr)
			continue

		if top_bitscore is None or bitscore > top_bitscore:
			top_bitscore = bitscore
			top_hit = cols

	return top_hit

def nearest_search(sample_genome, locus_list, represent_dir):
	# Check required external commands.
	check_command_exists("makeblastdb")
	check_command_exists("blastn")

	tophit_dic = {}
	# Create the BLAST database in a temporary directory.
	with tempfile.TemporaryDirectory() as tmpdir:
		db_prefix = Path(tmpdir) / "blastdb"
		make_blast_database(sample_genome, db_prefix)

		# Run blastn for each query FASTA file and print only the top hit.
		for locus in locus_list:
			blast_output = run_blastn(
				query_fasta=f"{represent_dir}/{locus}_represent.fasta",
				db_prefix=db_prefix
			)

			if blast_output == '':
				#print(f"{locus}\tNo Hit found")
				blast_output =  'No_Hit	No_Hit	0	0	0	0	0	0	0	0	0	0'
				#continue

			top_hit = get_top_hit(blast_output)
			
			tophit_dic[locus] = top_hit
			
	return(tophit_dic)
############################ Nearest allele search end ############################


def convert_to_phylocode(database_directory_path, cgmlst_df, genome_dic, represent_dir, outfile):
	# --------------------------
	# Step 1: Load all JSON conversion files from a directory
	# --------------------------
	print('Step 1: Load all Database files from database directory')
	json_dir = database_directory_path
	conversion_maps = {} # PhyloCode Database

	for file in os.listdir(json_dir):
		if file.endswith(".json"):
			locus = os.path.splitext(file)[0]
			with open(os.path.join(json_dir, file), 'r') as f:
				conversion_maps[locus] = json.load(f)

	# --------------------------
	# Step 2: Input allele DataFrame (with some missing values)
	# --------------------------
	print('Step 2: Input allele DataFrame (with some missing values)')
	allele_df = cgmlst_df
	# --------------------------
	# Step 3: Apply conversion using loaded maps
	# --------------------------
	print('Step 3: Apply conversion using loaded maps')
	converted_df = pd.DataFrame(index=allele_df.index, columns=allele_df.columns)
	
	outValue_dic = {}
	excelfile = f"{outfile}_alternate_allele.xlsx"
	for locus in allele_df.index:
		conv_map = conversion_maps.get(locus, {})  # If no JSON exists, empty dict
		for sample in allele_df.columns:
			original_value = allele_df.loc[locus, sample]
			if pd.isna(original_value):
				converted_value = -1
			else:
				try:
					allele_number = int(original_value)
				except ValueError:
					converted_value = -1
				else:
					if allele_number == -1:
						converted_value = -1
					else:
						key = f"{locus}_{allele_number}"
						converted_value = conv_map.get(key, -2)	  # <- Use -2 if key not found

			if converted_value == -2:
				if outValue_dic.get(sample) == None:
					outValue_dic[sample] =[locus]
				else:
					outValue_dic[sample] = outValue_dic[sample] + [locus]
				converted_value = -1
			converted_df.loc[locus, sample] = converted_value
	if len(outValue_dic) > 0 and len(genome_dic)>0 and represent_dir != "None":
		wb = openpyxl.Workbook()
		wb.save(excelfile)
		 
	for sample, loci in outValue_dic.items():
		if sample in genome_dic and represent_dir != "None":
			top_hit = nearest_search(genome_dic[sample], loci, represent_dir)
			#print(top_hit)

			# Column names corresponding to BLAST outfmt 6
			columns = [
				"allele", "alternate_allele", "sseqid", "pident", "length", "mismatch", "gapopen",
				"qstart", "qend", "sstart", "send", "evalue", "bitscore"
			]

			# Convert dictionary to DataFrame
			df = pd.DataFrame.from_dict(top_hit, orient="index", columns=columns)

			# Move dictionary keys into a normal column
			df.insert(0, "Locus", df.index)

			# Reset index
			df = df.reset_index(drop=True)

			# Convert numeric columns
			numeric_columns = [
				"pident", "length", "mismatch", "gapopen",
				"qstart", "qend", "sstart", "send", "evalue", "bitscore"
			]
			df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric)
			
			i = 0
			for locus in df["Locus"]:
				df.at[i, "allele"] = f"{locus}_{allele_df.at[locus, sample]}"
				print(sample, df.at[i, "allele"], df.at[i, "alternate_allele"])
				conv_map = conversion_maps.get(locus, {})
				converted_value = conv_map.get(df.at[i, "alternate_allele"], -1) 
				converted_df.loc[locus, sample] = converted_value
				i += 1

			# Save as Excel file with a specified sheet name
			with pd.ExcelWriter(excelfile, engine="openpyxl", mode="a") as writer:
				df.to_excel(writer, sheet_name=sample, index=False)
	if os.path.isfile(excelfile):
		wb = openpyxl.load_workbook(excelfile)
		sheet_names = wb.sheetnames
		#print(sheet_names)
		if "Sheet" in sheet_names:
			wb.remove(wb["Sheet"])
			wb.save(excelfile)
		wb.close()

	# --------------------------
	# Step 4: Result
	# --------------------------
	print('Step 4: Result')
	return(converted_df)

def calculate_distance_matrix(phylocode_df):
	# ---------- Distance matrix calculation ----------
	# Replace -1 with -3.14
	converted_df = phylocode_df.replace(-1, -3.14)

	# Fill missing values with -3.14 as well (if any NaNs remain)
	converted_df = converted_df.fillna(-3.14)

	# Compute distance matrix (sum of absolute differences column-wise)
	distance_matrix = pd.DataFrame(index=converted_df.columns, columns=converted_df.columns, dtype=float)

	for col1 in converted_df.columns:
		for col2 in converted_df.columns:
			distance = np.abs(converted_df[col1] - converted_df[col2]).sum()
			distance_matrix.loc[col1, col2] = distance

	# Return distance matrix
	print(distance_matrix)
	return(distance_matrix)	


def save_phylocode(cgMLST_locus_file, pubMLST_list, pubMLST_file, enterobase_file, outfile, database_directory_path, genome_dic, represent_dir):
	pubMLST_csv_list = []
	if pubMLST_list != 'None':
		with open(pubMLST_list, 'r') as file:
			for line in file:
				pubMLST_csv_list.append(line.strip())  # Add line to the list, stripping any leading/trailing whitespace

	if pubMLST_file != 'None':
		pubMLST_csv_list.append(pubMLST_file)

	if len(pubMLST_csv_list) == 0:
		pubMLST_csv_list = 'None'
	else:
		print(pubMLST_csv_list)

	# Read cgMLST profiles and convert to PhyloCode
	cgMLST_df = read_cgMLST(cgMLST_locus_file, pubMLST_csv_list, enterobase_file)
	print(cgMLST_df)
	phylocode_df = convert_to_phylocode(database_directory_path, cgMLST_df, genome_dic, represent_dir, outfile)
	print(phylocode_df)

	distance_matrix = calculate_distance_matrix(phylocode_df)

	# Save to CSV
	phylocode_df.to_csv(f'{outfile}_PhyloCode.csv', index=True)
	distance_matrix.to_csv(f'{outfile}_distance_matrix.csv', index=True)
	#cgMLST_df.to_csv(f'{outfile}_cgMLST.csv', index=True)
	return(phylocode_df)


# Convert radian values to RGB
def radian_to_color(rad_value, diff):
	x = rad_value
	if not (- np.pi <= x <= np.pi): #(0 <= x <= np.pi):
		if diff:
			R = 1
			G = 1
			B = 1
			#[1,1,1] == 'white'
		else:
			R = 0
			G = 0
			B = 0
			#[0,0,0] == 'black'
		#raise ValueError("Input must be in the range [0, pi]")
	elif diff:
		x = rad_value * 1.5
		if x >= np.pi:
			x = np.pi
		elif x<= -np.pi:
			x = -np.pi

		if x == 0:
			R = 0
			G = 0
			B = 0
		elif x < -(np.pi / 2):
			R = 0
			G = - np.sin(x)
			B = 1
		elif x >= -(np.pi / 2) and x < 0:
			R = 0
			G = int(255 - 201 * np.cos(x)**2) / 255
			B = -np.sin(x)**3
		elif x > 0 and x <= np.pi / 2:
			R = np.sin(x)
			G = int(255 - 201 * np.cos(x)**2) / 255
			B = 0
		elif x > np.pi / 2:
			R = 1
			G = np.sin(x)
			B = 0
		
	else:
		color = plt.cm.turbo(x / np.pi)
		R = color[0]
		G = color[1]
		B = color[2]
	return (R, G, B, 1)
	

def plot_circle_rings(radius_list, colors_list, labels_list, title, label_interval, file_name, diff):
	num_rings = len(radius_list)
	bar_distance = 0.5  # Distance between pointing bars and bars
	label_distance = 1.0  # Distance between labels and pointing bars
	
	fig, ax = plt.subplots(figsize=(16, 16))
	for i in range(num_rings):
		radius = radius_list[i]
		num_dots = len(colors_list[i])
		angle_step = 2 * np.pi / num_dots
		for j in range(num_dots):
			angle = np.pi / 2 - j * angle_step  # Start from 12 o'clock and move clockwise
			x_start = radius * np.cos(angle)
			y_start = radius * np.sin(angle)
			x_end = (radius - 0.5) * np.cos(angle)
			y_end = (radius - 0.5) * np.sin(angle)
			ax.plot([x_start, x_end], [y_start, y_end], color=colors_list[i][j % len(colors_list[i])], linewidth=2)
			if i == num_rings - 1 and j % label_interval == 0 and num_dots - j >=40:  # Label every 50th data point in the outer ring
				label_radius = radius + label_distance
				label_x = label_radius * np.cos(angle)
				label_y = label_radius * np.sin(angle)
				ax.text(label_x, label_y, labels_list[j % len(labels_list)], ha='center', va='center')
				# Add pointing bars
				bar_length = 0.5
				bar_radius = radius + bar_distance
				bar_x = bar_radius * np.cos(angle)
				bar_y = bar_radius * np.sin(angle)
				ax.plot([x_end, bar_x], [y_end, bar_y], color='black', linewidth=1)

	all_colors = []  # Collect all colors for the calibration bar
	if diff:
		all_colors = np.arange(-2.12, 2.12, 0.01)
	else:
		all_colors = np.arange(0, 3.12, 0.01)

	n_colors = len(all_colors)

	colorbar_ax = fig.add_axes([0.25, 0.03, 0.5, 0.04])  # [left, bottom, width, height]
	#cmap = plt.cm.rainbow

	colorbar_ax.set_xlim(0, n_colors)
	colorbar_ax.set_ylim(0, 1)
	colorbar_ax.axis('off')

	for i, color in enumerate(all_colors):
		rect = Rectangle((i, 0), 1, 1, color=radian_to_color(round(color,2), diff))
		colorbar_ax.add_patch(rect)

	# Numeric labels at both ends and center
	tick_fontsize = 9
	colorbar_ax.text(0, -0.3, str(round(all_colors[0],2)), ha='center', va='top', fontsize=tick_fontsize)
	colorbar_ax.text(n_colors // 2, -0.3, str(round(all_colors[n_colors // 2], 2)), ha='center', va='top', fontsize=tick_fontsize)
	colorbar_ax.text(n_colors - 1, -0.3, str(round(all_colors[n_colors - 1],2)), ha='center', va='top', fontsize=tick_fontsize)

	ax.set_aspect('equal')
	ax.set_title(title)
	ax.set_xticks([])
	ax.set_yticks([])
	ax.axis('off')

	plt.savefig(file_name)

def plot_linear_rings(radius_list, colors_list, labels_list, title, label_interval, ring_labels, file_name, diff):
	max_radius = max(radius_list)  # Determine the maximum radius
	bar_distance = 0.4  # Distance between pointing bars and bars
	label_distance = 0.5  # Distance between labels and pointing bars
	
	fig, ax = plt.subplots(figsize=(25, 16))
	for i in range(len(radius_list)):
		radius = max_radius  # Use max_radius for consistent width
		num_dots = len(colors_list[i])
		x_step = 2 * radius / num_dots
		x_start = -radius
		y_offset = -i * 0.8  # Vertical offset for each ring

		ax.text(x_start - 0.5, y_offset -0.25, ring_labels[i], ha='right', va='center', fontsize=10, fontweight='bold')
		
		for j in range(num_dots):
			x = x_start + j * x_step
			y_start = y_offset
			y_end = y_offset - 0.5
			ax.plot([x, x], [y_start, y_end], color=colors_list[i][j % len(colors_list[i])], linewidth=2)
			if i == len(radius_list) - 1 and j % label_interval == 0:  # Label every 50th data point in the outer ring
				label_y = y_end - label_distance
				ax.text(x, label_y, labels_list[j % len(labels_list)], ha='center', va='top', rotation=270)
				# Add pointing bars
				bar_y = y_end - bar_distance
				ax.plot([x, x], [y_end, bar_y], color='black', linewidth=1)


	all_colors = []  # Collect all colors for the calibration bar
	if diff:
		all_colors = np.arange(-2.12, 2.12, 0.01)
	else:
		all_colors = np.arange(0, 3.12, 0.01)

	n_colors = len(all_colors)

	colorbar_ax = fig.add_axes([0.25, 0.03, 0.5, 0.04])  # [left, bottom, width, height]
	#cmap = plt.cm.rainbow

	colorbar_ax.set_xlim(0, n_colors)
	colorbar_ax.set_ylim(0, 1)
	colorbar_ax.axis('off')

	for i, color in enumerate(all_colors):
		rect = Rectangle((i, 0), 1, 1, color=radian_to_color(round(color,2), diff))
		colorbar_ax.add_patch(rect)

	# Numeric labels at both ends and center
	tick_fontsize = 9
	colorbar_ax.text(0, -0.3, str(round(all_colors[0],2)), ha='center', va='top', fontsize=tick_fontsize)
	colorbar_ax.text(n_colors // 2, -0.3, str(round(all_colors[n_colors // 2], 2)), ha='center', va='top', fontsize=tick_fontsize)
	colorbar_ax.text(n_colors - 1, -0.3, str(round(all_colors[n_colors - 1],2)), ha='center', va='top', fontsize=tick_fontsize)

	ax.set_aspect('equal')
	ax.set_title(title)
	ax.set_xticks([])
	ax.set_yticks([])
	ax.axis('off')

	#plt.tight_layout()
	plt.savefig(file_name)
	

def adjust_code(phylocode_df, ref_strain, strain_list, fig_locus_list, remainref):
	# Subset first
	working_df = phylocode_df.copy()

	if len(strain_list) > 1:
		working_df = working_df[strain_list]

	if len(fig_locus_list) > 1:
		working_df = working_df.loc[fig_locus_list]

	# Missing or unmatched alleles are encoded as -1.
	missing_mask = working_df.eq(-1) | working_df.isna()

	if ref_strain != 'None':
		if ref_strain in working_df.columns:
			diff = True
		else:
			print(f"{ref_strain} does NOT exist in the cgMLST strains.")
			sys.exit()

		# Calculate differences using numeric values.
		diff_df = working_df.subtract(working_df[ref_strain], axis=0)

		# Any locus missing in either the sample or the reference should be non-comparable.
		ref_missing = missing_mask[ref_strain]

		# Expand the reference missing mask to the same shape as missing_mask.
		ref_missing_df = pd.DataFrame(
			{col: ref_missing for col in missing_mask.columns},
			index=missing_mask.index
		)

		non_comparable_mask = missing_mask | ref_missing_df

		# Use an out-of-range value so radian_to_color(..., diff=True) displays white.
		diff_df = diff_df.mask(non_comparable_mask, 100)
		if remainref:
			colormap_df = diff_df
		else:
			colormap_df = diff_df.drop(columns=[ref_strain])

		title = f"Difference color map using reference :{ref_strain}"

	else:
		diff = False

		# For ordinary PhyloCode maps, missing values are displayed as black.
		colormap_df = working_df.mask(missing_mask, 100)

		title = "Genome color map"

	return(colormap_df, title, diff)

def main():
	parser = argparse.ArgumentParser(description='color mapping')
	parser.add_argument(
		'-l', '--list',
		type = str,
		dest = 'pubMLST_Excel_list',
		required = False,
		default = 'None',
		help = 'pubMLST cgMLST results Excel file list'
	)
	parser.add_argument(
		'-p', '--pubmlst',
		type = str,
		dest = 'pubMLST_Excel_file',
		required = False,
		default = 'None',
		help = 'pubMLST cgMLST csv file'
	)
	parser.add_argument(
		'-e', '--enterobase',
		type = str,
		dest = 'enterobase_file',
		required = False,
		default = 'None',
		help = 'cgMLST profile obtained from Enterobase)'
	)
	parser.add_argument(
		'--phylocode',
		dest = 'PhyloCode',
		required = False,
		default = 'None',
		help = 'PhyloCode csv previously converted from cgMLST results.'
	) 
	parser.add_argument(
		'--genome',
		type = str,
		dest = 'genome_file',
		required = False,
		default = 'None',
		help = 'strain<TAB>genome.fasta genome fasta file to identify closely related allele if the allele does not exist in the database, --represent option is required.'
	)
	parser.add_argument(
		'--represent',
		type = str,
		dest = 'represent_alleles_dir',
		required = False,
		default = 'None',
		help = 'directory including representative alleles fasta'
	)
	parser.add_argument(
		'-c', '--locus',
		type = str,
		dest = 'locus_list_file',
		required = False,
		default = 'None',
		help = 'List of locus names, one per line. Each name must match a database file name (without extension). If a locus name in the cgMLST results differs from the list, append it to the list with a tab delimiter. Loci are sorted according to the order in the list.'
	)
	parser.add_argument(
		'-d', '--db',
		type = str,
		dest = 'database_dir_path',
		required = False,
		default = 'None',
		help = 'DATABASE path'
	)
	parser.add_argument(
		'-o', '--out',
		type = str,
		dest = 'out_file',
		default = 'compare_mapping',
		required = True,
		help = 'Prefix of output file'
	)
	parser.add_argument(
		'-m', '--map',
		dest = 'draw_map',
		action="store_true",
		default = False,
		help = 'Draw color map with linear mode'
	)
	parser.add_argument(
		'--circular',
		type = int,
		dest = 'max_num_circular',
		default = -1,
		help = 'Draw circular map and set the maximum number of strains. If the number of strain excess, draw with linear mode'
	)
	parser.add_argument(
		'-r', '--ref',
		type = str,
		dest = 'ref_strain',
		required = False,
		default = 'None',
		help = 'Reference strain name. Reference must be included in pubMLST or enterobase cgMLST file. Strain name for pubMLST results is the file name without extension.'
	)
	parser.add_argument(
		'--remainref',
		dest = 'remain_ref',
		action="store_true",
		default = False,
		help = 'Remain reference strain in the color map'
	)
	parser.add_argument(
		'--strain',
		type = str,
		dest = 'strain_list',
		required = False,
		default = 'None',
		help = 'List of strain to include color map, one per line.'
	)
	parser.add_argument(
		'--label_interval',
		type = int,
		dest = 'label_interval',
		default = 50,
		help = 'label_interval (default 50)'
	)
	parser.add_argument(
		'--open_viewer',
		dest = 'open_viewer',
		action="store_true",
		default = False,
		help = 'Open viewer'
	)

	args = parser.parse_args()

	pubMLST_list = args.pubMLST_Excel_list
	pubMLST_file = args.pubMLST_Excel_file
	enterobase_file = args.enterobase_file
	cgMLST_locus_file = args.locus_list_file
	database_directory_path = args.database_dir_path
	outfile = args.out_file
	phylocode_file = args.PhyloCode
	drawmap = args.draw_map
	circular = args.max_num_circular
	ref_strain = args.ref_strain
	remainref = args.remain_ref
	strain_list_file = args.strain_list
	label_interval = args.label_interval
	genome_file = args.genome_file
	represent_dir = args.represent_alleles_dir
	open_viewer = args.open_viewer


	# Option check
	if pubMLST_list != 'None' or pubMLST_file != "None" or enterobase_file != "None":
		if database_directory_path == 'None':
			print('-d (DATABASE path) is required.')
			sys.exit()
		elif cgMLST_locus_file == 'None':
			print('-c (cgMLST locus file) is required.')
			sys.exit()
	elif phylocode_file != 'None':
		pass
	else:
		print('Requirement: -l (pubMLST results list) or -p (pubMLST results file) or -e (Enterobase cgMLST) or --phylocode (Pre-converted PhyloCode file)')
		sys.exit()

	fig_locus_list = []
	if cgMLST_locus_file == 'None':
		pass
	else:
		with open(cgMLST_locus_file, 'r') as file:
			pre_locus_list = [line.strip() for line in file]
		for each in pre_locus_list:
			each_item = each.split()
			if len(each_item) == 1:
				fig_locus_list.append(each)
			else:
				fig_locus_list.append(each_item[0])
	#print(fig_locus_list)

	if strain_list_file == 'None':
		strain_list = []
	else:
		with open(strain_list_file, 'r') as file:
			strain_list = [line.strip() for line in file]
		if ref_strain in strain_list:
			pass
		elif ref_strain != 'None':
			strain_list.append(ref_strain)
		print(strain_list)
	
	genome_dic = {}
	if genome_file != 'None':
		with open(genome_file, 'r') as file:
			genome_list = [line.strip() for line in file]
		for line in genome_list:
			genome_dic[line.split()[0]] = line.split()[1]
		print(genome_dic)

	if pubMLST_list != 'None' or pubMLST_file != "None" or enterobase_file != "None":
		phylocode_df = save_phylocode(cgMLST_locus_file, pubMLST_list, pubMLST_file, enterobase_file, outfile, database_directory_path, genome_dic, represent_dir)
	elif phylocode_file != 'None':
		phylocode_df = pd.read_csv(phylocode_file, converters={0: str}, index_col=0)


	################################ Drawing color map ################################
	if drawmap == False and circular == -1:
		print('No color map drawing')
		sys.exit()

	adjusted_df, title, diff = adjust_code(phylocode_df, ref_strain, strain_list, fig_locus_list, remainref)
	strain_list = adjusted_df.columns.tolist()
	all_locus_list = adjusted_df.index.tolist()

	# Create color-mapped lists using plt.cm.rainbow
	final_color_lists = []

	for col in adjusted_df.columns:
		column_colors = [radian_to_color(value, diff) for value in adjusted_df[col]]  # RGBA tuples for this column
		final_color_lists.append(column_colors)

	# final_color_lists is a list of lists (one list per column of color-mapped values)
	# Example: Print number of column lists and first column's color values
	print(f"The number of strain is {len(final_color_lists)}")


	radius_list = [] 
	for i in range(len(final_color_lists)):
		radius_list.append(7.6 + 0.7 * i)

	file_name = f'{outfile}_map.eps'
	if drawmap:
		plot_linear_rings(radius_list, final_color_lists, all_locus_list, title, label_interval, strain_list, file_name, diff)
	elif len(final_color_lists) <= circular:
		plot_circle_rings(radius_list, final_color_lists, all_locus_list, title, label_interval, file_name, diff)
	elif circular == -1:
		pass
	elif len(final_color_lists) > circular:
		plot_linear_rings(radius_list, final_color_lists, all_locus_list, title, label_interval, strain_list, file_name, diff)
	if open_viewer:
		subprocess.Popen(['evince', file_name])

if __name__ == "__main__":
    main()

